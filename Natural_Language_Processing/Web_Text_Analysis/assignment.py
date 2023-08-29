import os
import openpyxl
import requests
from bs4 import BeautifulSoup
import nltk
from nltk.tokenize import sent_tokenize, word_tokenize
from nltk.corpus import cmudict, stopwords
import glob
import string
import re

nltk.download('punkt')
nltk.download('cmudict')
nltk.download('stopwords')


def extract_article_text(url):
    try:
        response = requests.get(url)
        soup = BeautifulSoup(response.content, 'html.parser')
        article_title = soup.find('h1').text.strip()
        article_text = '\n'.join(p.text.strip() for p in soup.find_all('p'))
        return article_title, article_text
    except Exception as e:
        print(f"Error while extracting article from {url}: {e}")
        return None, None


def create_directory(directory_name):
    if not os.path.exists(directory_name):
        os.makedirs(directory_name)


def process_excel_file(input_file, output_directory):
    workbook = openpyxl.load_workbook(input_file)
    sheet = workbook.active

    url_to_text = {}

    for row in sheet.iter_rows(min_row=2, values_only=True):
        url_id, url = int(row[0]), row[1]
        if url:
            article_title, article_text = extract_article_text(url)
            if article_text:
                url_to_text[url_id] = (url, article_text)  # Store url_id as the key in the dictionary

    for url_id, (url, article_text) in url_to_text.items():  # Unpack the url_id and url from the dictionary
        create_directory(output_directory)
        output_filename = os.path.join(output_directory, f"{url_id}.txt")
        with open(output_filename, 'w', encoding='utf-8') as output_file:
            output_file.write(article_text)

    return url_to_text


def load_stopwords(stopwords_files):
    stopwords_set = set()
    for file in stopwords_files:
        try:
            with open(file, 'r', encoding='utf-8') as f:
                stopwords_set.update(f.read().splitlines())
        except UnicodeDecodeError:
            with open(file, 'r', encoding='latin-1') as f:
                stopwords_set.update(f.read().splitlines())

    return stopwords_set


def create_positive_negative_dictionary(master_dictionary_folder, stopwords_set):
    positive_words = set()
    negative_words = set()

    master_dictionary_files = glob.glob(os.path.join(master_dictionary_folder, "*.txt"))

    for file in master_dictionary_files:
        try:
            with open(file, 'r', encoding='utf-8') as f:
                for line in f:
                    word = line.strip().lower()
                    if word not in stopwords_set:
                        if "positive" in file:
                            positive_words.add(word)
                        elif "negative" in file:
                            negative_words.add(word)
        except UnicodeDecodeError:
            with open(file, 'r', encoding='latin-1') as f:
                for line in f:
                    word = line.strip().lower()
                    if word not in stopwords_set:
                        if "positive" in file:
                            positive_words.add(word)
                        elif "negative" in file:
                            negative_words.add(word)

    return positive_words, negative_words


def calculate_derived_variables(text, positive_words, negative_words):
    tokens = word_tokenize(text)
    positive_score = sum(1 for word in tokens if word.lower() in positive_words)
    negative_score = -sum(-1 for word in tokens if word.lower() in negative_words)
    polarity_score = (positive_score - negative_score) / (positive_score + negative_score + 0.000001)
    subjectivity_score = (positive_score + negative_score) / (len(tokens) + 0.000001)

    return positive_score, negative_score, polarity_score, subjectivity_score


def count_complex_words(text):
    d = cmudict.dict()
    words = word_tokenize(text)
    complex_word_count = 0

    for word in words:
        if word.lower() not in d:
            continue
        syllables = [len(list(y for y in x if y[-1].isdigit())) for x in d[word.lower()]]
        if max(syllables) >= 3:
            complex_word_count += 1

    return complex_word_count


def calculate_gunning_fox_index(text):
    # Tokenize the text into sentences
    sentences = sent_tokenize(text)

    # Average Sentence Length
    total_words = len(word_tokenize(text))
    average_sentence_length = total_words / len(sentences)

    # Number of complex words
    complex_word_count = count_complex_words(text)

    # Percentage of Complex words
    percentage_complex_words = complex_word_count / total_words

    # Gunning Fox Index
    fog_index = 0.4 * (average_sentence_length + 100 * percentage_complex_words)

    return average_sentence_length, percentage_complex_words, fog_index


def calculate_average_words_per_sentence(text):
    # Tokenize the text into sentences
    sentences = sent_tokenize(text)

    # Total number of words
    total_words = len(word_tokenize(text))

    # Total number of sentences
    total_sentences = len(sentences)

    # Average number of words per sentence
    average_words_per_sentence = total_words / total_sentences

    return average_words_per_sentence


def calculate_word_count(text):
    # Tokenize the text into words
    words = word_tokenize(text)

    # Remove punctuation from words
    table = str.maketrans('', '', string.punctuation)
    words_without_punct = [word.translate(table) for word in words]

    # Remove stop words
    stop_words = set(nltk.corpus.stopwords.words('english'))
    cleaned_words = [word for word in words_without_punct if word.lower() not in stop_words]

    # Calculate the word count
    word_count = len(cleaned_words)

    return word_count


def count_syllables_per_word(word):
    # Count the number of syllables in the word
    vowels = "AEIOUYaeiouy"
    syllable_count = 0
    last_letter = None

    for letter in word:
        if letter in vowels and (last_letter is None or last_letter not in vowels):
            syllable_count += 1
        last_letter = letter

    # Handle exceptions for words ending with "es" or "ed"
    if word.endswith("es") or word.endswith("ed"):
        syllable_count -= 1

    return syllable_count


def count_personal_pronouns(text):
    # Define the list of personal pronouns
    personal_pronouns = ["I", "we", "my", "ours", "us"]

    # Define the regex pattern to match the personal pronouns
    pronoun_pattern = r'\b(?:' + '|'.join(personal_pronouns) + r')\b'

    # Find all occurrences of the personal pronouns in the text
    pronoun_matches = re.findall(pronoun_pattern, text, re.IGNORECASE)

    # Filter out the matches that are the country name "US"
    pronoun_count = len([pronoun for pronoun in pronoun_matches if pronoun.lower() != "us"])

    return pronoun_count


def calculate_average_word_length(text):
    # Tokenize the text into words
    words = word_tokenize(text)

    # Sum of total characters in each word
    total_characters = sum(len(word) for word in words)

    # Total number of words
    total_words = len(words)

    # Average word length
    average_word_length = total_characters / total_words

    return average_word_length


if __name__ == "__main__":
    input_file = "Input.xlsx"
    output_directory = "data_extracted"
    stopwords_directory = r"StopWords"
    stopwords_files = ["StopWords_Auditor.txt", "StopWords_Currencies.txt", "StopWords_DatesandNumbers.txt",
                       "StopWords_Generic.txt", "StopWords_GenericLong.txt", "StopWords_Geographic.txt",
                       "StopWords_Names.txt"]  # Add your list of stopwords files here
    master_dictionary_folder = r"MasterDictionary"

    if os.path.exists(input_file):
        # Get the URL_ID and URL to article text mapping
        url_to_text = process_excel_file(input_file, output_directory)
    else:
        print("Input file not found.")

    stopwords_files = [os.path.join(stopwords_directory, file) for file in stopwords_files]
    stopwords = load_stopwords(stopwords_files)

    # Dictionary of positive and negative words
    positive_words, negative_words = create_positive_negative_dictionary(master_dictionary_folder, stopwords)

    # Create an output Excel workbook
    output_workbook = openpyxl.Workbook()
    output_sheet = output_workbook.active

    # Set the headers for the columns in the output sheet
    output_sheet.append([
        "URL_ID", "URL", "POSITIVE SCORE", "NEGATIVE SCORE", "POLARITY SCORE", "SUBJECTIVITY SCORE",
        "AVG SENTENCE LENGTH", "PERCENTAGE OF COMPLEX WORDS", "FOG INDEX",
        "AVG NUMBER OF WORDS PER SENTENCE", "COMPLEX WORD COUNT", "WORD COUNT",
        "SYLLABLE PER WORD", "PERSONAL PRONOUNS", "AVG WORD LENGTH"])

    # Perform sentiment analysis on each extracted text file
    for url_id, (url, text) in url_to_text.items():
        if text is None:
            # Handle the case when the article text extraction failed
            continue

        positive_score, negative_score, polarity_score, subjectivity_score = calculate_derived_variables(text,
                                                                                                         positive_words,
                                                                                                         negative_words)
        # Calculate the average sentence length, percentage complex words and
        # Gunning Fox Index (Readability Analysis)
        average_sentence_length, percentage_complex_words, fog_index = calculate_gunning_fox_index(text)

        # Calculate the average number of words per sentence
        average_words_per_sentence = calculate_average_words_per_sentence(text)

        # Calculate the complex word count
        complex_word_count = count_complex_words(text)

        # Calculate the word count after removing stop words and punctuation
        word_count = calculate_word_count(text)

        # Calculate the syllable count per word
        words = word_tokenize(text)
        syllables_per_word = [count_syllables_per_word(word) for word in words]
        syllables_per_word = sum(syllables_per_word) / len(syllables_per_word)

        # Calculate the count of personal pronouns in the text
        personal_pronoun_count = count_personal_pronouns(text)

        # Calculate the average word length in the text
        average_word_length = calculate_average_word_length(text)

        positive_score = round(positive_score, 2)
        negative_score = round(negative_score, 2)
        polarity_score = round(polarity_score, 2)
        subjectivity_score = round(subjectivity_score, 2)
        average_sentence_length = round(average_sentence_length, 2)
        percentage_complex_words = round(percentage_complex_words, 2)
        fog_index = round(fog_index, 2)
        average_words_per_sentence = round(average_words_per_sentence, 2)
        complex_word_count = round(complex_word_count, 2)
        word_count = round(word_count, 2)
        syllables_per_word = round(syllables_per_word, 2)
        average_word_length = round(average_word_length, 2)

        output_sheet.append([
            url_id, url, positive_score, negative_score, polarity_score, subjectivity_score,
            average_sentence_length, percentage_complex_words, fog_index,
            average_words_per_sentence, complex_word_count, word_count,
            syllables_per_word, personal_pronoun_count, average_word_length])

    output_workbook.save("Output.xlsx")
    print("Analysis completed. Results saved to 'Output.xlsx'")
